import openpyxl
from openpyxl.styles import PatternFill, Alignment
from datetime import timedelta
import os
from pathlib import Path

def get_desktop_path():
    home = str(Path.home())
    paths = [
        os.path.join(home, "OneDrive", "デスクトップ"),
        os.path.join(home, "OneDrive", "Desktop"),
        os.path.join(home, "デスクトップ"),
        os.path.join(home, "Desktop")
    ]
    for p in paths:
        if os.path.exists(p):
            return p
    return os.getcwd()

def generate_list_excel(tasks, file_path=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks List"
    
    headers = ["ID", "Task Name", "Start Date", "End Date", "Progress (%)", "Assignee", "Milestone", "Dependencies"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        
    for row, t in enumerate(tasks, 2):
        ws.cell(row=row, column=1, value=t.id)
        ws.cell(row=row, column=2, value=t.name)
        ws.cell(row=row, column=3, value=t.start_date.strftime("%Y/%m/%d") if t.start_date else "")
        ws.cell(row=row, column=4, value=t.end_date.strftime("%Y/%m/%d") if t.end_date else "")
        ws.cell(row=row, column=5, value=t.progress)
        ws.cell(row=row, column=6, value=t.assignee)
        ws.cell(row=row, column=7, value="Yes" if t.milestone else "No")
        ws.cell(row=row, column=8, value=t.dependencies)
        
    if not file_path:
        file_path = os.path.join(get_desktop_path(), "tasks_list.xlsx")
    wb.save(file_path)
    return file_path

def generate_gantt_excel(tasks, file_path=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"
    
    if not tasks:
        ws.cell(row=1, column=1, value="No tasks available")
        if not file_path:
            file_path = os.path.join(get_desktop_path(), "gantt_chart.xlsx")
        wb.save(file_path)
        return file_path

    # Determine date range
    start_dates = [t.start_date for t in tasks if t.start_date]
    end_dates = [t.end_date for t in tasks if t.end_date]
    
    if not start_dates or not end_dates:
         ws.cell(row=1, column=1, value="Invalid dates")
         if not file_path:
            file_path = os.path.join(get_desktop_path(), "gantt_chart.xlsx")
         wb.save(file_path)
         return file_path

    min_date = min(start_dates)
    max_date = max(end_dates)
    
    # Headers
    ws.cell(row=1, column=1, value="タスク名")
    ws.cell(row=1, column=2, value="担当者")
    ws.cell(row=1, column=3, value="工数(h)")
    ws.cell(row=1, column=4, value="開始日")
    ws.cell(row=1, column=5, value="終了日")
    ws.cell(row=1, column=6, value="進捗(%)")
    
    current_date = min_date
    col_idx = 7
    date_to_col = {}
    while current_date <= max_date:
        cell = ws.cell(row=1, column=col_idx, value=current_date.strftime("%m/%d"))
        # Remove vertical rotation
        # cell.alignment = Alignment(text_rotation=90)
        date_to_col[current_date] = col_idx
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 6 # Slightly wider for horizontal dates
        current_date += timedelta(days=1)
        col_idx += 1
        
    fill_color = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    
    for row, t in enumerate(tasks, 2):
        ws.cell(row=row, column=1, value=t.name)
        ws.cell(row=row, column=2, value=t.assignee or "")
        ws.cell(row=row, column=3, value=t.man_hours or 0)
        ws.cell(row=row, column=4, value=t.start_date.strftime("%Y/%m/%d") if t.start_date else "")
        ws.cell(row=row, column=5, value=t.end_date.strftime("%Y/%m/%d") if t.end_date else "")
        ws.cell(row=row, column=6, value=t.progress or 0)
        
        if t.start_date and t.end_date:
            curr = t.start_date
            while curr <= t.end_date:
                if curr in date_to_col:
                    ws.cell(row=row, column=date_to_col[curr]).fill = fill_color
                curr += timedelta(days=1)
                
    if not file_path:
        file_path = os.path.join(get_desktop_path(), "gantt_chart.xlsx")
    wb.save(file_path)
    return file_path
